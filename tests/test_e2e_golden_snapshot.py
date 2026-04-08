import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.generate_inventory_risk_report import generate_report_for_system


SNAPSHOT_DIR = Path(__file__).parent / "snapshots"


def _normalize_records(df: pd.DataFrame, cols: list[str]) -> list[dict]:
    out = []
    for row in df[cols].to_dict(orient="records"):
        normalized = {}
        for key, value in row.items():
            if pd.isna(value):
                normalized[key] = None
            elif isinstance(value, float):
                normalized[key] = round(value, 3)
            else:
                normalized[key] = value
        out.append(normalized)
    return out


def _write_excel(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False)


def _read_overview_group(output_file: Path, group_name: str) -> pd.DataFrame:
    overview = pd.read_excel(output_file, sheet_name="运行总览", header=1)
    return overview.loc[overview["分组"] == group_name].copy()


def _capture_report(output_file: Path, expected_snapshot_name: str) -> None:
    usage_guide = pd.read_excel(output_file, sheet_name="使用说明", header=1)
    detail = pd.read_excel(output_file, sheet_name="明细", header=1)
    detail["门店名称"] = detail["门店名称"].ffill()
    catalog = pd.read_excel(output_file, sheet_name="商品编码对照清单", header=1)
    replenish = pd.read_excel(output_file, sheet_name="建议补货清单", header=1)
    transfer = pd.read_excel(output_file, sheet_name="建议调货清单", header=1)
    summary = _read_overview_group(output_file, "核心指标")

    actual = {
        "sheet_names": pd.ExcelFile(output_file).sheet_names,
        "usage_guide": _normalize_records(
            usage_guide,
            ["模块", "说明", "使用建议"],
        ),
        "detail": _normalize_records(
            detail,
            ["门店名称", "商品名称", "商品条码", "近三月+本月迄今平均日销", "库存数量", "风险等级", "建议调出数量", "建议补货数量"],
        ),
        "catalog": _normalize_records(
            catalog,
            ["商品编码", "品牌", "标准商品名", "销售表商品名", "库存商品名", "来源状态"],
        ),
        "replenish": _normalize_records(
            replenish,
            [c for c in ["门店名称", "商品名称", "省份", "装箱数（因子）", "建议补货数量", "建议补货箱数"] if c in replenish.columns],
        ),
        "transfer": _normalize_records(
            transfer,
            [c for c in ["门店名称", "商品名称", "省份", "装箱数（因子）", "建议调出数量"] if c in transfer.columns],
        ),
        "summary": {k: float(v) for k, v in zip(summary["指标"], summary["数值"])},
    }

    expected_path = SNAPSHOT_DIR / expected_snapshot_name
    with expected_path.open("r", encoding="utf-8") as f:
        expected = json.load(f)

    assert actual == expected


def test_golden_snapshot_shaanxi_baseline(tmp_path: Path):
    raw_root = tmp_path / "raw_data"
    system_dir = raw_root / "陕西华润"
    data_dir = tmp_path / "data"
    reports_dir = tmp_path / "reports"

    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A", "门店A"],
                "品牌": ["品牌X", "品牌X"],
                "商品名称": ["SKU高库存", "SKU缺货"],
                "国条码": ["6900000000001", "6900000000002"],
                "销售数量": [10, 20],
                "销售时间": ["2026-01-20", "2026-01-20"],
            }
        ),
        system_dir / "销售202601.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A", "门店A"],
                "品牌": ["品牌X", "品牌X"],
                "商品名称": ["SKU高库存", "SKU缺货"],
                "国条码": ["6900000000001", "6900000000002"],
                "销售数量": [15, 30],
                "销售时间": ["2026-02-08", "2026-02-08"],
            }
        ),
        system_dir / "销售202602.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A", "门店A"],
                "品牌": ["品牌X", "品牌X"],
                "商品名称": ["SKU高库存", "SKU缺货"],
                "国条码": ["6900000000001", "6900000000002"],
                "库存数量": [300, 0],
                "库存日期": ["2026-02-09", "2026-02-09"],
            }
        ),
        system_dir / "库存.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "商品条码": ["6900000000001", "6900000000002"],
                "商品名称": ["SKU高库存", "SKU缺货"],
                "装箱数（因子）": [10, 12],
            }
        ),
        data_dir / "sku装箱数.xlsx",
    )

    output_file = reports_dir / "陕西华润20260209库存预警.xlsx"
    config = {
        "run_mode": "single",
        "system_id": "shaanxi_huarun",
        "display_name": "陕西华润",
        "raw_data_dir": str(raw_root),
        "data_subdir": "陕西华润",
        "sales_files": ["销售202601.xlsx", "销售202602.xlsx"],
        "inventory_file": "库存.xlsx",
        "output_file": str(output_file),
        "carton_factor_file": str(data_dir / "sku装箱数.xlsx"),
        "risk_days_high": 60,
        "risk_days_low": 45,
        "sales_window_full_months": 3,
        "sales_window_include_mtd": True,
        "sales_window_recent_days": 30,
        "season_mode": False,
        "strict_auto_scan": False,
        "brand_keywords": ["品牌X"],
        "sales_date_dayfirst": False,
        "sales_date_format": "",
        "fail_on_empty_window": False,
    }
    generate_report_for_system(config, config)
    _capture_report(output_file, "golden_report_snapshot.json")


def test_golden_snapshot_wumei_province_and_national_barcode(tmp_path: Path):
    raw_root = tmp_path / "raw_data"
    system_dir = raw_root / "宁夏物美"
    data_dir = tmp_path / "data"
    reports_dir = tmp_path / "reports"

    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["物美A店"],
                "品牌": ["品牌Y"],
                "商品名称": ["物美SKU"],
                "商品编码": ["123456"],
                "国条码": ["6977777777777"],
                "销售数量": [20],
                "销售时间": ["2026-02-08"],
                "供商卡号": ["153085"],
            }
        ),
        system_dir / "销售202602.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["物美A店"],
                "品牌": ["品牌Y"],
                "商品名称": ["物美SKU"],
                "商品编码": ["123456"],
                "当前库存": [0],
                "库存日期": ["2026-02-09"],
                "供商卡号": ["153085"],
            }
        ),
        system_dir / "库存.xls",
    )
    _write_excel(
        pd.DataFrame(
            {
                "商品条码": ["6977777777777"],
                "商品名称": ["物美SKU"],
                "装箱数（因子）": [8],
            }
        ),
        data_dir / "sku装箱数.xlsx",
    )

    output_file = reports_dir / "宁夏物美20260209库存预警.xlsx"
    config = {
        "run_mode": "single",
        "system_id": "ningxia_wumei",
        "display_name": "宁夏物美",
        "raw_data_dir": str(raw_root),
        "data_subdir": "宁夏物美",
        "sales_files": ["销售202602.xlsx"],
        "inventory_file": "库存.xls",
        "output_file": str(output_file),
        "carton_factor_file": str(data_dir / "sku装箱数.xlsx"),
        "risk_days_high": 60,
        "risk_days_low": 45,
        "sales_window_full_months": 3,
        "sales_window_include_mtd": True,
        "sales_window_recent_days": 30,
        "season_mode": False,
        "strict_auto_scan": False,
        "brand_keywords": ["品牌Y"],
        "sales_date_dayfirst": False,
        "sales_date_format": "",
        "fail_on_empty_window": False,
    }
    generate_report_for_system(config, config)
    _capture_report(output_file, "golden_report_snapshot_wumei.json")


def test_golden_snapshot_peak_mode_case_rounding(tmp_path: Path):
    raw_root = tmp_path / "raw_data"
    system_dir = raw_root / "甘肃华润"
    data_dir = tmp_path / "data"
    reports_dir = tmp_path / "reports"

    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店B"],
                "品牌": ["品牌Z"],
                "商品名称": ["旺季SKU"],
                "国条码": ["6999999999999"],
                "销售数量": [9],
                "销售时间": ["2026-01-15"],
            }
        ),
        system_dir / "销售202601.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店B"],
                "品牌": ["品牌Z"],
                "商品名称": ["旺季SKU"],
                "国条码": ["6999999999999"],
                "销售数量": [30],
                "销售时间": ["2026-02-08"],
            }
        ),
        system_dir / "销售202602.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店B"],
                "品牌": ["品牌Z"],
                "商品名称": ["旺季SKU"],
                "国条码": ["6999999999999"],
                "库存数量": [0],
                "库存日期": ["2026-02-09"],
            }
        ),
        system_dir / "库存.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "商品条码": ["6999999999999"],
                "商品名称": ["旺季SKU"],
                "装箱数（因子）": [7],
            }
        ),
        data_dir / "sku装箱数.xlsx",
    )

    output_file = reports_dir / "甘肃华润20260209库存预警.xlsx"
    config = {
        "run_mode": "single",
        "system_id": "gansu_huarun",
        "display_name": "甘肃华润",
        "raw_data_dir": str(raw_root),
        "data_subdir": "甘肃华润",
        "sales_files": ["销售202601.xlsx", "销售202602.xlsx"],
        "inventory_file": "库存.xlsx",
        "output_file": str(output_file),
        "carton_factor_file": str(data_dir / "sku装箱数.xlsx"),
        "risk_days_high": 60,
        "risk_days_low": 45,
        "sales_window_full_months": 3,
        "sales_window_include_mtd": True,
        "sales_window_recent_days": 30,
        "season_mode": True,
        "strict_auto_scan": False,
        "brand_keywords": ["品牌Z"],
        "sales_date_dayfirst": False,
        "sales_date_format": "",
        "fail_on_empty_window": False,
    }
    generate_report_for_system(config, config)
    _capture_report(output_file, "golden_report_snapshot_peak.json")


def test_multi_store_product_code_matching_uses_store_code_and_does_not_fallback_by_product_name(tmp_path: Path):
    raw_root = tmp_path / "raw_data"
    system_dir = raw_root / "陕西华润"
    data_dir = tmp_path / "data"
    reports_dir = tmp_path / "reports"

    _write_excel(
        pd.DataFrame(
            {
                "门店编码": ["101", "202"],
                "门店名称": ["门店A", "门店B"],
                "品牌": ["品牌F", "品牌F"],
                "商品名称": ["同名SKU", "同名SKU"],
                "商品编码": ["P-A", "P-B"],
                "商品条码": ["S-A", "S-B"],
                "销售数量": [30, 6],
                "销售时间": ["2026-02-08", "2026-02-08"],
            }
        ),
        system_dir / "销售202602.xlsx",
    )
    _write_excel(
        pd.DataFrame(
                {
                    "门店编码": ["101", "202"],
                    "门店名称": ["门店A", "门店B"],
                    "品牌": ["品牌F", "品牌F"],
                    "商品名称": ["同名SKU", "同名SKU"],
                    "商品编码": ["P-A", "P-B"],
                    "商品编码.1": ["MISMATCH-A", "MISMATCH-B"],
                    "库存数量": [100, 100],
                    "库存日期": ["2026-02-09", "2026-02-09"],
                }
        ),
        system_dir / "库存.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "商品条码": [None],
                "商品名称": ["同名SKU"],
                "装箱数（因子）": [10],
            }
        ),
        data_dir / "sku装箱数.xlsx",
    )

    output_file = reports_dir / "陕西华润20260209库存预警.xlsx"
    config = {
        "run_mode": "single",
        "system_id": "shaanxi_huarun",
        "display_name": "陕西华润",
        "raw_data_dir": str(raw_root),
        "data_subdir": "陕西华润",
        "sales_files": ["销售202602.xlsx"],
        "inventory_file": "库存.xlsx",
        "output_file": str(output_file),
        "carton_factor_file": str(data_dir / "sku装箱数.xlsx"),
        "risk_days_high": 60,
        "risk_days_low": 45,
        "sales_window_full_months": 3,
        "sales_window_include_mtd": True,
        "sales_window_recent_days": 30,
        "season_mode": False,
        "strict_auto_scan": False,
        "brand_keywords": ["品牌F"],
        "sales_date_dayfirst": False,
        "sales_date_format": "",
        "fail_on_empty_window": False,
    }

    generate_report_for_system(config, config)
    detail = pd.read_excel(output_file, sheet_name="明细", header=1)
    detail["门店名称"] = detail["门店名称"].ffill()
    a_daily = float(detail.loc[detail["门店名称"] == "门店A", "近三月+本月迄今平均日销"].iloc[0])
    b_daily = float(detail.loc[detail["门店名称"] == "门店B", "近三月+本月迄今平均日销"].iloc[0])
    assert a_daily > 0.0
    assert b_daily > 0.0


def test_wumei_missing_national_barcode_and_supplier_fallbacks(tmp_path: Path):
    raw_root = tmp_path / "raw_data"
    system_dir = raw_root / "宁夏物美"
    data_dir = tmp_path / "data"
    reports_dir = tmp_path / "reports"

    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["物美B店"],
                "品牌": ["品牌M"],
                "商品名称": ["无国条码SKU"],
                "商品编码": ["817620"],
                "销售数量": [10],
                "销售时间": ["2026-02-08"],
            }
        ),
        system_dir / "销售202602.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["物美B店"],
                "品牌": ["品牌M"],
                "商品名称": ["无国条码SKU"],
                "商品编码": ["817620"],
                "当前库存": [0],
                "库存日期": ["2026-02-09"],
            }
        ),
        system_dir / "库存.xls",
    )
    _write_excel(
        pd.DataFrame(
            {
                "商品条码": ["817620"],
                "商品名称": ["无国条码SKU"],
                "装箱数（因子）": [6],
            }
        ),
        data_dir / "sku装箱数.xlsx",
    )

    output_file = reports_dir / "宁夏物美20260209库存预警.xlsx"
    config = {
        "run_mode": "single",
        "system_id": "ningxia_wumei",
        "display_name": "宁夏物美",
        "raw_data_dir": str(raw_root),
        "data_subdir": "宁夏物美",
        "sales_files": ["销售202602.xlsx"],
        "inventory_file": "库存.xls",
        "output_file": str(output_file),
        "carton_factor_file": str(data_dir / "sku装箱数.xlsx"),
        "risk_days_high": 60,
        "risk_days_low": 45,
        "sales_window_full_months": 3,
        "sales_window_include_mtd": True,
        "sales_window_recent_days": 30,
        "season_mode": False,
        "strict_auto_scan": False,
        "brand_keywords": ["品牌M"],
        "sales_date_dayfirst": False,
        "sales_date_format": "",
        "fail_on_empty_window": False,
    }

    generate_report_for_system(config, config)
    detail = pd.read_excel(output_file, sheet_name="明细", header=1)
    detail["门店名称"] = detail["门店名称"].ffill()
    row = detail.iloc[0]
    assert str(row["商品条码"]) == "817620"
    assert row["省份"] == "其他/未知"


def test_zero_sales_stagnant_inventory_generates_outbound_qty(tmp_path: Path):
    raw_root = tmp_path / "raw_data"
    system_dir = raw_root / "陕西华润"
    data_dir = tmp_path / "data"
    reports_dir = tmp_path / "reports"

    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A"],
                "品牌": ["品牌S"],
                "商品名称": ["滞销SKU"],
                "国条码": ["6901111111111"],
                "销售数量": [0],
                "销售时间": ["2026-02-08"],
            }
        ),
        system_dir / "销售202602.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A"],
                "品牌": ["品牌S"],
                "商品名称": ["滞销SKU"],
                "国条码": ["6901111111111"],
                "库存数量": [20],
                "库存日期": ["2026-02-09"],
            }
        ),
        system_dir / "库存.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "商品条码": ["6901111111111"],
                "商品名称": ["滞销SKU"],
                "装箱数（因子）": [6],
            }
        ),
        data_dir / "sku装箱数.xlsx",
    )

    output_file = reports_dir / "陕西华润20260209库存预警.xlsx"
    config = {
        "run_mode": "single",
        "system_id": "shaanxi_huarun",
        "display_name": "陕西华润",
        "raw_data_dir": str(raw_root),
        "data_subdir": "陕西华润",
        "sales_files": ["销售202602.xlsx"],
        "inventory_file": "库存.xlsx",
        "output_file": str(output_file),
        "carton_factor_file": str(data_dir / "sku装箱数.xlsx"),
        "risk_days_high": 60,
        "risk_days_low": 45,
        "sales_window_full_months": 3,
        "sales_window_include_mtd": True,
        "sales_window_recent_days": 30,
        "season_mode": False,
        "strict_auto_scan": False,
        "stagnant_min_keep_qty": 2,
        "brand_keywords": ["品牌S"],
        "sales_date_dayfirst": False,
        "sales_date_format": "",
        "fail_on_empty_window": False,
    }

    generate_report_for_system(config, config)
    transfer = pd.read_excel(output_file, sheet_name="建议调货清单", header=1)
    assert len(transfer) == 1
    assert int(transfer["建议调出数量"].iloc[0]) == 18


def test_zero_sales_stagnant_inventory_supports_all_outbound_mode(tmp_path: Path):
    raw_root = tmp_path / "raw_data"
    system_dir = raw_root / "陕西华润"
    data_dir = tmp_path / "data"
    reports_dir = tmp_path / "reports"

    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A"],
                "品牌": ["品牌S"],
                "商品名称": ["滞销SKU"],
                "国条码": ["6901111111111"],
                "销售数量": [0],
                "销售时间": ["2026-02-08"],
            }
        ),
        system_dir / "销售202602.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A"],
                "品牌": ["品牌S"],
                "商品名称": ["滞销SKU"],
                "国条码": ["6901111111111"],
                "库存数量": [20],
                "库存日期": ["2026-02-09"],
            }
        ),
        system_dir / "库存.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "商品条码": ["6901111111111"],
                "商品名称": ["滞销SKU"],
                "装箱数（因子）": [6],
            }
        ),
        data_dir / "sku装箱数.xlsx",
    )

    output_file = reports_dir / "陕西华润20260209库存预警.xlsx"
    config = {
        "run_mode": "single",
        "system_id": "shaanxi_huarun",
        "display_name": "陕西华润",
        "raw_data_dir": str(raw_root),
        "data_subdir": "陕西华润",
        "sales_files": ["销售202602.xlsx"],
        "inventory_file": "库存.xlsx",
        "output_file": str(output_file),
        "carton_factor_file": str(data_dir / "sku装箱数.xlsx"),
        "risk_days_high": 60,
        "risk_days_low": 45,
        "sales_window_full_months": 3,
        "sales_window_include_mtd": True,
        "sales_window_recent_days": 30,
        "season_mode": False,
        "strict_auto_scan": False,
        "stagnant_outbound_mode": "all_outbound",
        "stagnant_min_keep_qty": 5,
        "brand_keywords": ["品牌S"],
        "sales_date_dayfirst": False,
        "sales_date_format": "",
        "fail_on_empty_window": False,
    }

    generate_report_for_system(config, config)
    transfer = pd.read_excel(output_file, sheet_name="建议调货清单", header=1)
    assert len(transfer) == 1
    assert int(transfer["建议调出数量"].iloc[0]) == 20


def test_detail_sheet_can_disable_store_name_merge_cells(tmp_path: Path):
    raw_root = tmp_path / "raw_data"
    system_dir = raw_root / "陕西华润"
    data_dir = tmp_path / "data"
    reports_dir = tmp_path / "reports"

    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A", "门店A"],
                "品牌": ["品牌T", "品牌T"],
                "商品名称": ["SKU1", "SKU2"],
                "国条码": ["6902111111111", "6902111111112"],
                "销售数量": [10, 8],
                "销售时间": ["2026-02-08", "2026-02-08"],
            }
        ),
        system_dir / "销售202602.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A", "门店A"],
                "品牌": ["品牌T", "品牌T"],
                "商品名称": ["SKU1", "SKU2"],
                "国条码": ["6902111111111", "6902111111112"],
                "库存数量": [100, 80],
                "库存日期": ["2026-02-09", "2026-02-09"],
            }
        ),
        system_dir / "库存.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "商品条码": ["6902111111111", "6902111111112"],
                "商品名称": ["SKU1", "SKU2"],
                "装箱数（因子）": [6, 6],
            }
        ),
        data_dir / "sku装箱数.xlsx",
    )

    output_file = reports_dir / "陕西华润20260209库存预警.xlsx"
    config = {
        "run_mode": "single",
        "system_id": "shaanxi_huarun",
        "display_name": "陕西华润",
        "raw_data_dir": str(raw_root),
        "data_subdir": "陕西华润",
        "sales_files": ["销售202602.xlsx"],
        "inventory_file": "库存.xlsx",
        "output_file": str(output_file),
        "carton_factor_file": str(data_dir / "sku装箱数.xlsx"),
        "risk_days_high": 60,
        "risk_days_low": 45,
        "sales_window_full_months": 3,
        "sales_window_include_mtd": True,
        "sales_window_recent_days": 30,
        "season_mode": False,
        "strict_auto_scan": False,
        "merge_detail_store_cells": False,
        "brand_keywords": ["品牌T"],
        "sales_date_dayfirst": False,
        "sales_date_format": "",
        "fail_on_empty_window": False,
    }

    generate_report_for_system(config, config)
    wb = load_workbook(output_file)
    detail_ws = wb["明细"]
    merged_ranges = {str(cell_range) for cell_range in detail_ws.merged_cells.ranges}
    assert "A3:A4" not in merged_ranges


def test_report_includes_store_sales_ranking_transfer_sheet(tmp_path: Path):
    raw_root = tmp_path / "raw_data"
    system_dir = raw_root / "陕西华润"
    data_dir = tmp_path / "data"
    reports_dir = tmp_path / "reports"

    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A", "门店A", "门店B"],
                "品牌": ["品牌T", "品牌T", "品牌T"],
                "商品名称": ["SKU1", "SKU2", "SKU1"],
                "商品条码": ["6902111111111", "6902111111112", "6902111111111"],
                "销售数量": [0, 10, 20],
                "销售金额": [100.0, 50.0, 300.0],
                "销售时间": ["2026-02-08", "2026-02-08", "2026-02-08"],
            }
        ),
        system_dir / "销售202602.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A", "门店A", "门店B"],
                "品牌": ["品牌T", "品牌T", "品牌T"],
                "商品名称": ["SKU1", "SKU2", "SKU1"],
                "商品条码": ["6902111111111", "6902111111112", "6902111111111"],
                "库存数量": [8, 20, 80],
                "库存日期": ["2026-02-09", "2026-02-09", "2026-02-09"],
            }
        ),
        system_dir / "库存.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "商品条码": ["6902111111111", "6902111111112"],
                "商品名称": ["SKU1", "SKU2"],
                "装箱数（因子）": [6, 6],
            }
        ),
        data_dir / "sku装箱数.xlsx",
    )

    output_file = reports_dir / "陕西华润20260209库存预警.xlsx"
    config = {
        "run_mode": "single",
        "system_id": "shaanxi_huarun",
        "display_name": "陕西华润",
        "raw_data_dir": str(raw_root),
        "data_subdir": "陕西华润",
        "sales_files": ["销售202602.xlsx"],
        "inventory_file": "库存.xlsx",
        "output_file": str(output_file),
        "carton_factor_file": str(data_dir / "sku装箱数.xlsx"),
        "risk_days_high": 60,
        "risk_days_low": 45,
        "sales_window_full_months": 3,
        "sales_window_include_mtd": True,
        "sales_window_recent_days": 30,
        "season_mode": False,
        "strict_auto_scan": False,
        "merge_detail_store_cells": True,
        "brand_keywords": ["品牌T"],
        "sales_date_dayfirst": False,
        "sales_date_format": "",
        "fail_on_empty_window": False,
    }

    generate_report_for_system(config, config)
    ranking_sheet = pd.read_excel(output_file, sheet_name="门店销量排名调货汇总", header=1)
    assert ranking_sheet["门店名称"].tolist() == ["门店B", "门店A"]
    assert ranking_sheet["排名"].tolist() == [1, 2]
    assert ranking_sheet["门店销售额总计"].tolist() == [300.0, 150.0]
    assert ranking_sheet["商品销售额"].tolist() == [300.0, 100.0]
    assert ranking_sheet["调货数量"].tolist() == [40, 8]


def test_sales_supplier_fallback_handles_duplicate_store_barcode(tmp_path: Path):
    raw_root = tmp_path / "raw_data"
    system_dir = raw_root / "陕西华润"
    data_dir = tmp_path / "data"
    reports_dir = tmp_path / "reports"

    _write_excel(
        pd.DataFrame(
            {
                "门店编码": ["101", "101"],
                "门店名称": ["门店A", "门店A"],
                "品牌": ["品牌Q", "品牌Q"],
                "商品名称": ["商品名1", "商品名2"],
                "商品编码": ["P-1", "P-1"],
                "商品条码": ["6904007562070", "6904007562070"],
                "销售数量": [5, 6],
                "销售时间": ["2026-02-08", "2026-02-08"],
            }
        ),
        system_dir / "销售202602.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "门店编码": ["101"],
                "门店名称": ["门店A"],
                "品牌": ["品牌Q"],
                "商品名称": ["库存商品X"],
                "商品编码": ["P-1"],
                "商品编码.1": ["6904007562070"],
                "库存数量": [10],
                "库存日期": ["2026-02-09"],
            }
        ),
        system_dir / "库存.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "商品条码": ["6904007562070"],
                "商品名称": ["库存商品X"],
                "装箱数（因子）": [6],
            }
        ),
        data_dir / "sku装箱数.xlsx",
    )

    output_file = reports_dir / "陕西华润20260209库存预警.xlsx"
    config = {
        "run_mode": "single",
        "system_id": "shaanxi_huarun",
        "display_name": "陕西华润",
        "raw_data_dir": str(raw_root),
        "data_subdir": "陕西华润",
        "sales_files": ["销售202602.xlsx"],
        "inventory_file": "库存.xlsx",
        "output_file": str(output_file),
        "carton_factor_file": str(data_dir / "sku装箱数.xlsx"),
        "risk_days_high": 60,
        "risk_days_low": 45,
        "sales_window_full_months": 3,
        "sales_window_include_mtd": True,
        "sales_window_recent_days": 30,
        "season_mode": False,
        "strict_auto_scan": False,
        "brand_keywords": ["品牌Q"],
        "sales_date_dayfirst": False,
        "sales_date_format": "",
        "fail_on_empty_window": False,
    }

    generate_report_for_system(config, config)
    assert output_file.exists()
    detail = pd.read_excel(output_file, sheet_name="明细", header=1)
    detail["门店名称"] = detail["门店名称"].ffill()
    rows = detail.loc[detail["门店名称"] == "门店A"]
    assert len(rows) == 1
    row = rows.iloc[0]
    assert float(row["近三月+本月迄今平均日销"]) > 0
    assert row["商品名称"] == "商品名2"

    status = _read_overview_group(output_file, "运行状态")
    status_map = dict(zip(status["指标"], status["数值"]))
    assert int(status_map["同店同商品编码重复键数"]) == 1
    assert int(status_map["名称冲突键数"]) == 1
    assert int(status_map["品牌冲突键数"]) == 0


def test_brand_is_derived_when_column_missing_or_blank(tmp_path: Path):
    raw_root = tmp_path / "raw_data"
    system_dir = raw_root / "陕西华润"
    data_dir = tmp_path / "data"
    reports_dir = tmp_path / "reports"

    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A", "门店A"],
                "商品名称": ["伊利高钙奶", "未知商品"],
                "商品条码": ["6901111111111", "6902222222222"],
                "销售数量": [10, 5],
                "销售时间": ["2026-02-08", "2026-02-08"],
            }
        ),
        system_dir / "销售202602.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "门店名称": ["门店A", "门店A"],
                "商品名称": ["伊利高钙奶", "未知商品"],
                "品牌": ["", None],
                "商品条码": ["6901111111111", "6902222222222"],
                "库存数量": [20, 10],
                "库存日期": ["2026-02-09", "2026-02-09"],
            }
        ),
        system_dir / "库存.xlsx",
    )
    _write_excel(
        pd.DataFrame(
            {
                "商品条码": ["6901111111111", "6902222222222"],
                "商品名称": ["伊利高钙奶", "未知商品"],
                "装箱数（因子）": [6, 6],
            }
        ),
        data_dir / "sku装箱数.xlsx",
    )

    output_file = reports_dir / "陕西华润20260209库存预警.xlsx"
    config = {
        "run_mode": "single",
        "system_id": "shaanxi_huarun",
        "display_name": "陕西华润",
        "raw_data_dir": str(raw_root),
        "data_subdir": "陕西华润",
        "sales_files": ["销售202602.xlsx"],
        "inventory_file": "库存.xlsx",
        "output_file": str(output_file),
        "carton_factor_file": str(data_dir / "sku装箱数.xlsx"),
        "risk_days_high": 60,
        "risk_days_low": 45,
        "sales_window_full_months": 3,
        "sales_window_include_mtd": True,
        "sales_window_recent_days": 30,
        "season_mode": False,
        "strict_auto_scan": False,
        "brand_keywords": ["伊利", "蒙牛"],
        "sales_date_dayfirst": False,
        "sales_date_format": "",
        "fail_on_empty_window": False,
    }
    generate_report_for_system(config, config)

    detail = pd.read_excel(output_file, sheet_name="明细", header=1)
    detail["门店名称"] = detail["门店名称"].ffill()
    brand_by_product = dict(zip(detail["商品名称"], detail["品牌"]))
    assert brand_by_product["伊利高钙奶"] == "伊利"
    assert brand_by_product["未知商品"] == "其他"
