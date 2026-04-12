from __future__ import annotations

from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import load_workbook

from . import report_styles as core_report_styles


def write_report_with_style(
    output_file: Path,
    display_name: str,
    inventory_date: str,
    sheets: Dict[str, pd.DataFrame],
    merge_detail_store_cells: bool = True,
) -> None:
    """写出 Excel 并统一应用工作表样式。"""
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_file)
    style_ctx = core_report_styles.build_style_context()
    center = style_ctx["center"]

    for sheet_name in sheets.keys():
        ws = wb[sheet_name]
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "1F4E79" if sheet_name == "运行总览" else "7B94B0"
        core_report_styles.set_sheet_title(ws, display_name, inventory_date, center)
        core_report_styles.set_freeze_and_filter(ws, sheet_name)
        headers = core_report_styles.style_headers(ws, style_ctx["header_fill"], style_ctx["header_font"], center)
        core_report_styles.apply_column_widths(ws)
        header_to_idx = core_report_styles.style_data_rows(
            ws,
            headers,
            style_ctx["left"],
            center,
            style_ctx["risk_fills"],
            style_ctx["out_of_stock_fill"],
            style_ctx["band_fill"],
            style_ctx["transfer_qty_fill"],
        )
        if sheet_name == "运行总览":
            core_report_styles.apply_overview_metric_format(ws, header_to_idx)

    ws_detail = wb["明细"] if "明细" in wb.sheetnames else None
    if ws_detail is not None and merge_detail_store_cells:
        core_report_styles.merge_detail_store_cells(ws_detail, center)

    core_report_styles.apply_borders(wb)
    wb.save(output_file)
