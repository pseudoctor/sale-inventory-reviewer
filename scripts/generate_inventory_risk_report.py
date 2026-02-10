#!/usr/bin/env python3
"""Generate inventory risk Excel report from windowed daily sales and inventory data."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from decimal import Decimal, InvalidOperation
from datetime import datetime
from copy import deepcopy
import re

import numpy as np
import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE_DIR = Path(__file__).parent.parent
CONFIG_PATH = BASE_DIR / "config.yaml"

DEFAULT_CONFIG = {
    "run_mode": "single",
    "display_name": "",
    "system_id": "",
    "raw_data_dir": "./raw_data",
    "output_file": "./reports/inventory_risk_report.xlsx",
    "sales_files": [],
    "inventory_file": "",
    "risk_days_high": 60,
    "risk_days_low": 45,
    "sales_window_full_months": 3,
    "sales_window_include_mtd": True,
    "sales_window_recent_days": 30,
    "sales_date_dayfirst": False,
    "sales_date_format": "",
    "season_mode": False,
    "fail_on_empty_window": False,
    "carton_factor_file": "./data/sku装箱数.xlsx",
    "brand_keywords": [],
    "batch": {
        "continue_on_error": True,
        "summary_output_file": "./reports/batch_run_summary.xlsx",
        "systems": [],
    },
}

PREFERRED_WIDTHS = {
    "门店名称": 18,
    "品牌": 10,
    "商品条码": 16,
    "商品名称": 36,
    "装箱数（因子）": 12,
    "近三月+本月迄今平均日销": 22,
    "近30天平均日销售": 16,
    "库存数量": 10,
    "风险等级": 8,
    "库存/销售比": 12,
    "库存周转率": 12,
    "库存周转天数": 12,
    "建议调出数量": 12,
    "建议补货数量": 12,
    "建议补货箱数": 12,
}

NUMBER_FORMATS = {
    "近三月+本月迄今平均日销": "0.000",
    "近30天平均日销售": "0.000",
    "库存数量": "0",
    "装箱数（因子）": "0",
    "库存/销售比": "0.0",
    "库存周转率": "0.0%",
    "库存周转天数": "0",
    "建议调出数量": "0",
    "建议补货数量": "0",
}

SUMMARY_ONE_DECIMAL_METRICS = {
    "近三月+本月迄今平均日销总量",
    "近30天平均日销售总量",
    "预测平均日销总量(季节模式后)",
}

DEFAULT_LEGACY_OUTPUT_FILES = {
    "./reports/inventory_risk_report.xlsx",
    "reports/inventory_risk_report.xlsx",
    "inventory_risk_report.xlsx",
}

AUTO_OUTPUT_DATE_PLACEHOLDER = "{库存日期}"


def validate_config(config: Dict[str, Any]) -> Dict[str, Any]:
    run_mode = str(config.get("run_mode", "single")).strip().lower()
    if run_mode not in {"single", "batch"}:
        raise ValueError("config.run_mode must be one of: single, batch.")
    config["run_mode"] = run_mode

    if not isinstance(config.get("raw_data_dir"), str) or not str(config["raw_data_dir"]).strip():
        raise ValueError("config.raw_data_dir must be a non-empty string.")
    display_name = config.get("display_name", "")
    if display_name is not None and not isinstance(display_name, str):
        raise ValueError("config.display_name must be a string.")
    system_id = config.get("system_id", "")
    if system_id is not None and not isinstance(system_id, str):
        raise ValueError("config.system_id must be a string.")
    if run_mode == "single":
        if not isinstance(config.get("output_file"), str) or not str(config["output_file"]).strip():
            raise ValueError("config.output_file must be a non-empty string in single mode.")
        if not isinstance(config.get("inventory_file"), str) or not str(config["inventory_file"]).strip():
            raise ValueError("config.inventory_file must be a non-empty string in single mode.")
    carton_factor_file = config.get("carton_factor_file", "")
    if not isinstance(carton_factor_file, str) or not carton_factor_file.strip():
        raise ValueError("config.carton_factor_file must be a non-empty string.")

    sales_files = config.get("sales_files")
    if not isinstance(sales_files, list):
        raise ValueError("config.sales_files must be a list.")
    if not all(isinstance(x, str) and x.strip() for x in sales_files):
        raise ValueError("config.sales_files entries must be non-empty strings.")

    risk_days_high = float(config.get("risk_days_high", 60))
    risk_days_low = float(config.get("risk_days_low", 45))
    if risk_days_high <= 0 or risk_days_low <= 0:
        raise ValueError("risk_days_high and risk_days_low must be positive.")
    if risk_days_low >= risk_days_high:
        raise ValueError("risk_days_low must be smaller than risk_days_high.")

    full_months = int(config.get("sales_window_full_months", 3))
    if full_months < 0:
        raise ValueError("sales_window_full_months must be >= 0.")
    recent_days = int(config.get("sales_window_recent_days", 30))
    if recent_days <= 0:
        raise ValueError("sales_window_recent_days must be > 0.")
    include_mtd = config.get("sales_window_include_mtd", True)
    if not isinstance(include_mtd, bool):
        raise ValueError("sales_window_include_mtd must be true/false.")
    sales_date_dayfirst = config.get("sales_date_dayfirst", False)
    if not isinstance(sales_date_dayfirst, bool):
        raise ValueError("sales_date_dayfirst must be true/false.")
    sales_date_format = config.get("sales_date_format", "")
    if not isinstance(sales_date_format, str):
        raise ValueError("sales_date_format must be a string.")

    season_mode = config.get("season_mode", False)
    if not isinstance(season_mode, (bool, str)):
        raise ValueError("season_mode must be true/false or legacy string peak/off_peak.")
    if isinstance(season_mode, str):
        mode_text = season_mode.strip().lower()
        if mode_text not in {"true", "false", "peak", "off_peak"}:
            raise ValueError("season_mode string must be one of: true, false, peak, off_peak.")
    fail_on_empty_window = config.get("fail_on_empty_window", False)
    if not isinstance(fail_on_empty_window, bool):
        raise ValueError("fail_on_empty_window must be true/false.")

    brand_keywords = config.get("brand_keywords")
    if not isinstance(brand_keywords, list):
        raise ValueError("brand_keywords must be a list.")
    if not all(isinstance(x, str) for x in brand_keywords):
        raise ValueError("brand_keywords entries must be strings.")

    batch = config.get("batch", {})
    if not isinstance(batch, dict):
        raise ValueError("config.batch must be a dict.")
    continue_on_error = batch.get("continue_on_error", True)
    if not isinstance(continue_on_error, bool):
        raise ValueError("config.batch.continue_on_error must be true/false.")
    summary_output_file = batch.get("summary_output_file", "./reports/batch_run_summary.xlsx")
    if not isinstance(summary_output_file, str) or not summary_output_file.strip():
        raise ValueError("config.batch.summary_output_file must be a non-empty string.")
    systems = batch.get("systems", [])
    if not isinstance(systems, list):
        raise ValueError("config.batch.systems must be a list.")
    batch["continue_on_error"] = continue_on_error
    batch["summary_output_file"] = summary_output_file
    batch["systems"] = systems
    config["batch"] = batch
    return config


def validate_batch_config(config: Dict[str, Any]) -> None:
    batch_cfg = config.get("batch", {})
    systems = batch_cfg.get("systems", [])
    if not systems:
        raise ValueError("batch mode requires at least one system in config.batch.systems.")

    seen_ids = set()
    seen_display_names = set()
    for idx, system in enumerate(systems, start=1):
        if not isinstance(system, dict):
            raise ValueError(f"batch.systems[{idx}] must be a dict.")
        enabled = system.get("enabled", True)
        if not isinstance(enabled, bool):
            raise ValueError(f"batch.systems[{idx}].enabled must be true/false.")
        system_id = str(system.get("system_id", "")).strip()
        display_name = system.get("display_name")
        if not isinstance(display_name, str) or not display_name.strip():
            raise ValueError(f"batch.systems[{idx}].display_name must be a non-empty string.")
        normalized_display_name = display_name.strip()
        if normalized_display_name in seen_display_names:
            raise ValueError(f"Duplicated display_name in batch.systems: {normalized_display_name}")
        seen_display_names.add(normalized_display_name)
        identity = system_id or display_name.strip()
        if identity in seen_ids:
            raise ValueError(f"Duplicated identity in batch.systems: {identity}")
        seen_ids.add(identity)
        data_subdir = system.get("data_subdir")
        if data_subdir is not None and (not isinstance(data_subdir, str) or not data_subdir.strip()):
            raise ValueError(f"batch.systems[{idx}].data_subdir must be a non-empty string if provided.")
        sales_files = system.get("sales_files")
        if not isinstance(sales_files, list) or not sales_files:
            raise ValueError(f"batch.systems[{idx}].sales_files must be a non-empty list.")
        if not all(isinstance(x, str) and x.strip() for x in sales_files):
            raise ValueError(f"batch.systems[{idx}].sales_files entries must be non-empty strings.")
        inventory_file = system.get("inventory_file")
        if not isinstance(inventory_file, str) or not inventory_file.strip():
            raise ValueError(f"batch.systems[{idx}].inventory_file must be a non-empty string.")
        output_file = system.get("output_file")
        if output_file is not None and (not isinstance(output_file, str) or not output_file.strip()):
            raise ValueError(f"batch.systems[{idx}].output_file must be a non-empty string if provided.")

    # Prevent accidental overwrite from duplicated explicit output path.
    seen_output_paths = set()
    for idx, system in enumerate(systems, start=1):
        output_file = system.get("output_file")
        if not isinstance(output_file, str) or not output_file.strip():
            continue
        path = Path(output_file.strip())
        normalized = str((BASE_DIR / path).resolve()) if not path.is_absolute() else str(path.resolve())
        if normalized in seen_output_paths:
            raise ValueError(f"Duplicated output_file in batch.systems[{idx}]: {output_file}")
        seen_output_paths.add(normalized)


def build_system_config(system_cfg: Dict[str, Any], global_cfg: Dict[str, Any]) -> Dict[str, Any]:
    merged = dict(global_cfg)
    merged["enabled"] = bool(system_cfg.get("enabled", True))
    merged["sales_files"] = list(system_cfg["sales_files"])
    merged["inventory_file"] = str(system_cfg["inventory_file"]).strip()
    merged["display_name"] = str(system_cfg["display_name"]).strip()
    raw_system_id = str(system_cfg.get("system_id", "")).strip()
    merged["system_id"] = raw_system_id or merged["display_name"]
    raw_data_subdir = str(system_cfg.get("data_subdir", "")).strip()
    merged["data_subdir"] = raw_data_subdir

    output_file = system_cfg.get("output_file")
    if isinstance(output_file, str) and output_file.strip():
        merged["output_file"] = output_file.strip()
    else:
        merged["output_file"] = ""

    carton_factor_file = system_cfg.get("carton_factor_file")
    if isinstance(carton_factor_file, str) and carton_factor_file.strip():
        merged["carton_factor_file"] = carton_factor_file.strip()

    return validate_config(merged)


def resolve_system_raw_data_dir(config: Dict[str, Any]) -> Path:
    raw_data_dir = Path(config["raw_data_dir"])
    if not raw_data_dir.is_absolute():
        raw_data_dir = (BASE_DIR / raw_data_dir).resolve()
    data_subdir = str(config.get("data_subdir", "")).strip()
    if data_subdir:
        raw_data_dir = (raw_data_dir / data_subdir).resolve()
    return raw_data_dir


def resolve_output_file_path(config: Dict[str, Any], display_name: str, inventory_date: str) -> Path:
    configured_output = str(config.get("output_file", "")).strip()
    inventory_date_compact = inventory_date.replace("-", "")
    if (
        not configured_output
        or configured_output in DEFAULT_LEGACY_OUTPUT_FILES
    ):
        configured_output = f"./reports/{display_name}{inventory_date_compact}库存预警.xlsx"
    output_file = Path(configured_output)
    if not output_file.is_absolute():
        output_file = (BASE_DIR / output_file).resolve()
    return output_file


def resolve_expected_output_for_status(config: Dict[str, Any], display_name: str) -> str:
    configured_output = str(config.get("output_file", "")).strip()
    if not configured_output or configured_output in DEFAULT_LEGACY_OUTPUT_FILES:
        configured_output = f"./reports/{display_name}{AUTO_OUTPUT_DATE_PLACEHOLDER}库存预警.xlsx"
    output_file = Path(configured_output)
    if not output_file.is_absolute():
        output_file = (BASE_DIR / output_file).resolve()
    return str(output_file)


def load_config() -> Dict[str, Any]:
    config = deepcopy(DEFAULT_CONFIG)
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            loaded = yaml.safe_load(f) or {}
        for key, value in loaded.items():
            if key == "batch" and isinstance(value, dict):
                config["batch"].update(value)
            else:
                config[key] = value
    return validate_config(config)


def extract_month_key(filename: str) -> Optional[int]:
    match = re.search(r"(\d{4})(\d{2})", filename)
    if not match:
        return None
    return int(match.group(1) + match.group(2))


def find_column(columns: List[str], candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in columns:
            return c
    return None


def resolve_sales_candidates(raw_data_dir: Path, configured_sales_files: List[str]) -> List[Path]:
    if configured_sales_files:
        return [raw_data_dir / name for name in configured_sales_files]

    candidates: List[Tuple[int, Path]] = []
    for path in raw_data_dir.glob("*.xlsx"):
        month_key = extract_month_key(path.name)
        if month_key is None:
            continue
        candidates.append((month_key, path))
    candidates.sort(key=lambda x: x[0])
    return [path for _, path in candidates]


def normalize_sales_df(df: pd.DataFrame) -> Tuple[pd.DataFrame, str, str, str, str, str, str]:
    df = df.copy()
    df.columns = df.columns.str.strip()

    store_col = find_column(df.columns, ["门店名称", "门店", "store"])
    brand_col = find_column(df.columns, ["品牌", "brand"])
    product_col = find_column(df.columns, ["商品名称", "商品", "product"])
    barcode_col = find_column(df.columns, ["商品条码", "条码", "商品编码.1", "商品编码", "barcode"])
    qty_col = find_column(df.columns, ["销售数量", "数量", "sales_qty", "qty"])
    date_col = find_column(df.columns, ["销售时间", "日期", "date", "sales_date"])

    if not all([store_col, brand_col, product_col, barcode_col, qty_col, date_col]):
        missing = [
            name
            for name, col in [
                ("门店名称", store_col),
                ("品牌", brand_col),
                ("商品名称", product_col),
                ("商品条码", barcode_col),
                ("销售数量", qty_col),
                ("销售时间", date_col),
            ]
            if col is None
        ]
        raise ValueError(f"Missing required sales columns: {', '.join(missing)}")

    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    return df, store_col, brand_col, product_col, barcode_col, qty_col, date_col


def overlap_days(
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
    data_min_date: pd.Timestamp,
    data_max_date: pd.Timestamp,
) -> int:
    overlap_start = max(start_date, data_min_date)
    overlap_end = min(end_date, data_max_date)
    if overlap_start > overlap_end:
        return 0
    return (overlap_end - overlap_start).days + 1


def combine_daily_sales(
    daily_sales_3m_mtd: pd.Series,
    daily_sales_30d: pd.Series,
    use_peak_mode: bool,
) -> pd.Series:
    if use_peak_mode:
        return pd.concat([daily_sales_3m_mtd, daily_sales_30d], axis=1).max(axis=1)
    return pd.concat([daily_sales_3m_mtd, daily_sales_30d], axis=1).min(axis=1)


def classify_risk_levels(turnover_days: pd.Series, low_days: float, high_days: float) -> pd.Series:
    return pd.Series(
        np.select(
            [turnover_days > high_days, turnover_days < low_days],
            ["高", "低"],
            default="中",
        ),
        index=turnover_days.index,
    )


def normalize_inventory_df(df: pd.DataFrame) -> Tuple[pd.DataFrame, str, Optional[str], str, str, str]:
    df = df.copy()
    df.columns = df.columns.str.strip()

    store_col = find_column(df.columns, ["门店名称", "门店", "store"])
    brand_col = find_column(df.columns, ["品牌", "brand"])
    product_col = find_column(df.columns, ["商品名称", "商品", "product"])
    barcode_col = find_column(df.columns, ["商品条码", "条码", "商品编码.1", "商品编码", "barcode"])
    qty_col = find_column(df.columns, ["数量", "库存数量", "inventory_qty", "qty"])

    if not all([store_col, product_col, barcode_col, qty_col]):
        missing = [
            name
            for name, col in [
                ("门店名称", store_col),
                ("商品名称", product_col),
                ("商品条码", barcode_col),
                ("数量", qty_col),
            ]
            if col is None
        ]
        raise ValueError(f"Missing required inventory columns: {', '.join(missing)}")

    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)

    return df, store_col, brand_col, product_col, barcode_col, qty_col


def normalize_barcode_value(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float):
        if pd.isna(value):
            return None
        return format(value, ".0f")
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if text == "" or text.lower() in {"nan", "none"}:
        return None
    # Normalize scientific notation text, e.g. "6.907992633671E12"
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?[eE][+-]?\d+", text):
        try:
            text = format(Decimal(text), "f")
        except (InvalidOperation, ValueError):
            return text
    # Normalize common spreadsheet artifacts such as "6907...0.0"
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    if re.fullmatch(r"\d+\.\d+", text):
        integer, decimal = text.split(".", 1)
        if set(decimal) == {"0"}:
            return integer
    return text


def parse_sales_dates(raw_dates: pd.Series, date_format: str, dayfirst: bool) -> pd.Series:
    if date_format.strip():
        parsed = pd.to_datetime(raw_dates, format=date_format.strip(), errors="coerce")
    else:
        parsed = pd.to_datetime(raw_dates, errors="coerce", dayfirst=dayfirst)
    return parsed


def compute_case_counts(qty: pd.Series, factor: pd.Series, use_peak_mode: bool) -> pd.Series:
    qty_num = pd.to_numeric(qty, errors="coerce").fillna(0)
    factor_num = pd.to_numeric(factor, errors="coerce")
    valid_factor = factor_num > 0
    raw_cases = np.where(valid_factor, qty_num / factor_num, np.nan)
    rounded = np.where(valid_factor, np.ceil(raw_cases) if use_peak_mode else np.floor(raw_cases), np.nan)
    return pd.Series(rounded, index=qty.index).astype("Int64")


def load_carton_factor_df(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Carton factor file not found: {path}")
    df = pd.read_excel(path, sheet_name=0, dtype=str)
    df.columns = df.columns.astype(str).str.strip()

    barcode_col = find_column(df.columns.tolist(), ["商品条码", "条码", "商品编码.1", "商品编码", "barcode"])
    product_col = find_column(df.columns.tolist(), ["商品名称", "商品", "product"])
    factor_col = find_column(df.columns.tolist(), ["装箱数（因子）", "装箱数(因子)", "装箱数", "因子", "factor"])
    if not all([barcode_col, product_col, factor_col]):
        missing = [
            name
            for name, col in [("商品条码", barcode_col), ("商品名称", product_col), ("装箱数（因子）", factor_col)]
            if col is None
        ]
        raise ValueError(f"Carton factor file missing columns: {', '.join(missing)}")

    out = df[[barcode_col, product_col, factor_col]].copy()
    out.columns = ["商品条码", "商品名称", "装箱数（因子）"]
    out["商品条码"] = out["商品条码"].apply(normalize_barcode_value)
    out["商品名称"] = out["商品名称"].astype(str).str.strip()
    out["装箱数（因子）"] = pd.to_numeric(out["装箱数（因子）"], errors="coerce")
    out = out[out["装箱数（因子）"].notna() & (out["装箱数（因子）"] > 0)]
    out["装箱数（因子）"] = out["装箱数（因子）"].astype(int)
    return out


def extract_brand_from_product(product: Optional[str], brands: List[str]) -> str:
    if product is None:
        return "其他"
    text = str(product)
    for brand in brands:
        if brand in text:
            return brand
    return "其他"


def ensure_inventory_brand_column(df: pd.DataFrame, brands: List[str]) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.str.strip()
    if "品牌" in df.columns:
        return df
    if "商品名称" not in df.columns:
        raise ValueError("Inventory file missing 商品名称; cannot derive 品牌 column.")
    brand_series = df["商品名称"].apply(lambda v: extract_brand_from_product(v, brands))
    insert_at = list(df.columns).index("商品名称")
    df.insert(insert_at, "品牌", brand_series)
    return df


def extract_inventory_date(df: pd.DataFrame) -> str:
    df = df.copy()
    df.columns = df.columns.str.strip()

    date_col = find_column(df.columns, ["库存日期", "日期", "盘点日期", "库存时间", "时间", "inventory_date", "date"])
    if not date_col:
        return "未知"

    raw_series = df[date_col]
    parsed = pd.to_datetime(raw_series, errors="coerce").dropna()
    if not parsed.empty:
        date_value = parsed.max().date()
        return date_value.isoformat()

    non_null = raw_series.dropna()
    if non_null.empty:
        return "未知"

    return str(non_null.iloc[0]).strip()


def apply_inventory_metrics(df: pd.DataFrame, low_days: float, high_days: float) -> pd.DataFrame:
    out = df.copy()
    out["inventory_sales_ratio"] = np.where(
        out["forecast_daily_sales"] > 0,
        out["inventory_qty"] / out["forecast_daily_sales"],
        float("inf"),
    )
    out["turnover_rate"] = np.where(
        out["inventory_qty"] > 0,
        (out["forecast_daily_sales"] * 30) / out["inventory_qty"],
        0,
    )
    turnover_precise = np.where(
        out["forecast_daily_sales"] > 0,
        out["inventory_qty"] / out["forecast_daily_sales"],
        float("inf"),
    )
    out["turnover_days"] = np.where(np.isfinite(turnover_precise), np.round(turnover_precise), float("inf"))
    out["risk_level"] = classify_risk_levels(pd.Series(turnover_precise, index=out.index), low_days, high_days)
    return out


def generate_report_for_system(system_cfg: Dict[str, Any], global_cfg: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    config = dict(system_cfg)
    _ = global_cfg
    system_id = str(config.get("system_id", "single")).strip() or "single"
    display_name = str(config.get("display_name", system_id)).strip() or system_id

    raw_data_dir = resolve_system_raw_data_dir(config)

    configured_sales_files = [f for f in config.get("sales_files") or []]
    inventory_file = config.get("inventory_file") or ""
    carton_factor_file = config.get("carton_factor_file") or ""
    brand_keywords = [str(b).strip() for b in (config.get("brand_keywords") or []) if str(b).strip()]
    if not brand_keywords:
        brand_keywords = [
            "伊利",
            "蒙牛",
            "南国",
            "阿宝乐",
            "畅鲜选",
            "蜀珑珠",
            "永璞",
            "雅士利",
            "大漠银根",
            "谷栗村",
            "君乐宝",
            "秦俑",
            "红色拖拉机",
            "飞鹤",
            "雀巢",
            "Seesaw",
            "佳贝艾特",
        ]

    sales_candidates = resolve_sales_candidates(raw_data_dir, configured_sales_files)

    # Inventory data
    inv_path = raw_data_dir / inventory_file
    if not inv_path.exists():
        raise FileNotFoundError(f"Inventory file not found: {inv_path}")

    inv_df = pd.read_excel(inv_path, sheet_name=0, dtype=str)
    inv_df = ensure_inventory_brand_column(inv_df, brand_keywords)
    inventory_date_text = extract_inventory_date(inv_df)
    parsed_inventory_date = pd.to_datetime(inventory_date_text, errors="coerce")
    if pd.isna(parsed_inventory_date):
        raise ValueError(f"Invalid inventory date: {inventory_date_text}")
    inventory_date_ts = pd.Timestamp(parsed_inventory_date).normalize()
    inventory_date = inventory_date_ts.date().isoformat()
    output_file = resolve_output_file_path(config, display_name, inventory_date)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    inv_df, inv_store, inv_brand, inv_product, inv_barcode, inv_qty = normalize_inventory_df(inv_df)
    if inv_brand is None:
        inv_df = inv_df.copy()
        inv_df["品牌"] = inv_df[inv_product].apply(lambda v: extract_brand_from_product(v, brand_keywords))
        inv_brand = "品牌"
    inv_df = inv_df[[inv_store, inv_brand, inv_product, inv_barcode, inv_qty]].copy()
    inv_df.columns = ["store", "brand", "product", "barcode", "inventory_qty"]
    inv_df["barcode"] = inv_df["barcode"].apply(normalize_barcode_value)

    full_months = int(config.get("sales_window_full_months", 3))
    include_mtd = bool(config.get("sales_window_include_mtd", True))
    recent_days = int(config.get("sales_window_recent_days", 30))
    sales_date_dayfirst = bool(config.get("sales_date_dayfirst", False))
    sales_date_format = str(config.get("sales_date_format", ""))
    season_mode_raw = config.get("season_mode", False)
    fail_on_empty_window = bool(config.get("fail_on_empty_window", False))
    if isinstance(season_mode_raw, bool):
        use_peak_mode = season_mode_raw
    else:
        mode_text = str(season_mode_raw).strip().lower()
        if mode_text in {"true", "peak"}:
            use_peak_mode = True
        elif mode_text in {"false", "off_peak"}:
            use_peak_mode = False
        else:
            raise ValueError("season_mode must be true/false (or legacy peak/off_peak)")
    if full_months < 0:
        raise ValueError("sales_window_full_months must be >= 0")
    if recent_days <= 0:
        raise ValueError("sales_window_recent_days must be > 0")

    sales_data = []
    invalid_sales_date_rows = 0
    for filepath in sales_candidates:
        if not filepath.exists():
            print(f"Warning: missing sales file {filepath}")
            continue
        df = pd.read_excel(filepath, sheet_name=0, dtype=str)
        df, store_col, brand_col, product_col, barcode_col, qty_col, date_col = normalize_sales_df(df)
        df = df[[store_col, brand_col, product_col, barcode_col, qty_col, date_col]].copy()
        df.columns = ["store", "brand", "product", "barcode", "sales_qty", "sales_date"]
        df["barcode"] = df["barcode"].apply(normalize_barcode_value)
        parsed_dates = parse_sales_dates(
            df["sales_date"],
            date_format=sales_date_format,
            dayfirst=sales_date_dayfirst,
        )
        invalid_sales_date_rows += int(parsed_dates.isna().sum())
        df["sales_date"] = parsed_dates
        df = df[df["sales_date"].notna()].copy()
        sales_data.append(df)

    if not sales_data:
        raise FileNotFoundError("No sales files were loaded.")

    carton_factor_path = Path(carton_factor_file)
    if not carton_factor_path.is_absolute():
        carton_factor_path = (BASE_DIR / carton_factor_path).resolve()
    carton_factor_df = load_carton_factor_df(carton_factor_path)
    sales_df = pd.concat(sales_data, ignore_index=True)

    mtd_end = inventory_date_ts if include_mtd else (inventory_date_ts.replace(day=1) - pd.Timedelta(days=1))
    mtd_start = (inventory_date_ts.replace(day=1) - pd.DateOffset(months=full_months)).normalize()
    recent_start = (inventory_date_ts - pd.Timedelta(days=recent_days - 1)).normalize()

    if sales_df.empty:
        raise ValueError("Sales data has no valid dates after parsing 销售时间.")

    data_min_date = sales_df["sales_date"].min().normalize()
    data_max_date = sales_df["sales_date"].max().normalize()
    mtd_days = overlap_days(mtd_start, mtd_end, data_min_date, data_max_date)
    recent_days_effective = overlap_days(recent_start, inventory_date_ts, data_min_date, data_max_date)
    has_mtd_window_data = mtd_days > 0
    has_recent_window_data = recent_days_effective > 0
    if fail_on_empty_window and (not has_mtd_window_data or not has_recent_window_data):
        raise ValueError(
            f"Sales window has no overlapping data: 3M+MTD={has_mtd_window_data}, 30D={has_recent_window_data}"
        )

    sales_mtd = sales_df[(sales_df["sales_date"] >= mtd_start) & (sales_df["sales_date"] <= mtd_end)]
    sales_30d = sales_df[(sales_df["sales_date"] >= recent_start) & (sales_df["sales_date"] <= inventory_date_ts)]

    sales_totals = (
        sales_df.groupby(["store", "brand", "product", "barcode"], as_index=False)["sales_qty"]
        .sum()
        .rename(columns={"sales_qty": "sales_qty_total"})
    )
    daily_3m_mtd = (
        sales_mtd.groupby(["store", "brand", "product", "barcode"], as_index=False)["sales_qty"]
        .sum()
        .rename(columns={"sales_qty": "sales_qty_3m_mtd"})
    )
    daily_30d = (
        sales_30d.groupby(["store", "brand", "product", "barcode"], as_index=False)["sales_qty"]
        .sum()
        .rename(columns={"sales_qty": "sales_qty_30d"})
    )
    sales_totals = sales_totals.merge(daily_3m_mtd, on=["store", "brand", "product", "barcode"], how="left")
    sales_totals = sales_totals.merge(daily_30d, on=["store", "brand", "product", "barcode"], how="left")
    sales_totals["sales_qty_3m_mtd"] = sales_totals["sales_qty_3m_mtd"].fillna(0)
    sales_totals["sales_qty_30d"] = sales_totals["sales_qty_30d"].fillna(0)
    sales_totals["daily_sales_3m_mtd"] = (
        sales_totals["sales_qty_3m_mtd"] / mtd_days if has_mtd_window_data else 0
    )
    sales_totals["daily_sales_30d"] = (
        sales_totals["sales_qty_30d"] / recent_days_effective if has_recent_window_data else 0
    )
    sales_totals["forecast_daily_sales"] = combine_daily_sales(
        sales_totals["daily_sales_3m_mtd"],
        sales_totals["daily_sales_30d"],
        use_peak_mode,
    )
    sales_totals["barcode_key"] = sales_totals["barcode"].apply(normalize_barcode_value)

    inv_totals = (
        inv_df.groupby(["store", "brand", "product", "barcode"], as_index=False)["inventory_qty"]
        .sum()
    )
    inv_totals["barcode_key"] = inv_totals["barcode"].apply(normalize_barcode_value)

    # Match sales to inventory: barcode match first, fallback to store+brand+product
    sales_barcode = sales_totals[sales_totals["barcode_key"].notna()].copy()

    detail = inv_totals.copy()
    detail = detail.merge(
        sales_barcode[["store", "barcode_key", "daily_sales_3m_mtd", "daily_sales_30d", "forecast_daily_sales"]],
        on=["store", "barcode_key"],
        how="left",
    )

    fallback_sales = sales_totals.groupby(["store", "brand", "product"], as_index=False).agg({
        "daily_sales_3m_mtd": "sum",
        "daily_sales_30d": "sum",
        "forecast_daily_sales": "sum",
    })
    missing_mask = detail["forecast_daily_sales"].isna()
    if missing_mask.any():
        fallback = detail.loc[missing_mask].merge(
            fallback_sales,
            on=["store", "brand", "product"],
            how="left",
            suffixes=("", "_fallback"),
        )
        detail.loc[missing_mask, "daily_sales_3m_mtd"] = fallback["daily_sales_3m_mtd_fallback"].values
        detail.loc[missing_mask, "daily_sales_30d"] = fallback["daily_sales_30d_fallback"].values
        detail.loc[missing_mask, "forecast_daily_sales"] = fallback["forecast_daily_sales_fallback"].values

    detail["daily_sales_3m_mtd"] = detail["daily_sales_3m_mtd"].fillna(0)
    detail["daily_sales_30d"] = detail["daily_sales_30d"].fillna(0)
    detail["forecast_daily_sales"] = detail["forecast_daily_sales"].fillna(0)

    high_days = float(config.get("risk_days_high", 60))
    low_days = float(config.get("risk_days_low", 45))
    detail = apply_inventory_metrics(detail, low_days, high_days)
    
    detail = detail[[
        "store",
        "brand",
        "barcode",
        "product",
        "daily_sales_3m_mtd",
        "daily_sales_30d",
        "forecast_daily_sales",
        "inventory_qty",
        "risk_level",
        "inventory_sales_ratio",
        "turnover_rate",
        "turnover_days",
    ]]

    detail = detail.sort_values(["store", "brand", "product", "barcode"]).reset_index(drop=True)

    # Store summary
    store_summary = detail.groupby("store", as_index=False).agg({
        "daily_sales_3m_mtd": "sum",
        "daily_sales_30d": "sum",
        "forecast_daily_sales": "sum",
        "inventory_qty": "sum",
    })
    store_summary = apply_inventory_metrics(store_summary, low_days, high_days)

    # Brand summary
    brand_summary = detail.groupby("brand", as_index=False).agg({
        "daily_sales_3m_mtd": "sum",
        "daily_sales_30d": "sum",
        "forecast_daily_sales": "sum",
        "inventory_qty": "sum",
    })
    brand_summary = apply_inventory_metrics(brand_summary, low_days, high_days)

    # Missing-in-inventory list: sales with qty but no inventory match
    inv_barcode_keys = set(zip(inv_totals["store"], inv_totals["barcode_key"]))
    inv_fallback_keys = set(zip(inv_totals["store"], inv_totals["brand"], inv_totals["product"]))

    def is_missing(row) -> bool:
        if row["barcode_key"] is not None:
            if (row["store"], row["barcode_key"]) in inv_barcode_keys:
                return False
        if (row["store"], row["brand"], row["product"]) in inv_fallback_keys:
            return False
        return True

    missing_sales = sales_totals.copy()
    missing_sales["is_missing"] = missing_sales.apply(is_missing, axis=1)
    missing_sales = missing_sales[missing_sales["is_missing"] & (missing_sales["forecast_daily_sales"] > 0)]

    # Append missing items into detail table
    if not missing_sales.empty:
        missing_detail = missing_sales.copy()
        missing_detail["inventory_qty"] = 0
        missing_detail["inventory_sales_ratio"] = float("inf")
        missing_detail["turnover_rate"] = 0
        missing_detail["turnover_days"] = float("inf")
        missing_detail["risk_level"] = "高"
        missing_detail = missing_detail[[
            "store",
            "brand",
            "barcode",
            "product",
            "daily_sales_3m_mtd",
            "daily_sales_30d",
            "forecast_daily_sales",
            "inventory_qty",
                "risk_level",
            "inventory_sales_ratio",
            "turnover_rate",
            "turnover_days",
        ]]
        detail = pd.concat([detail, missing_detail], ignore_index=True)

    detail = detail.sort_values(["store", "brand", "product", "barcode"]).reset_index(drop=True)

    detail["out_of_stock"] = np.where(
        (detail["forecast_daily_sales"] > 0) & (detail["inventory_qty"] == 0),
        "是",
        "否",
    )

    # Suggested outbound/replenishment quantities (store-SKU only, no lane matching)
    detail["daily_demand"] = detail["forecast_daily_sales"]
    detail["low_target_qty"] = detail["daily_demand"] * low_days
    detail["high_keep_qty"] = detail["daily_demand"] * high_days
    detail["need_qty"] = np.where(
        detail["forecast_daily_sales"] > 0,
        np.ceil(np.maximum(0, detail["low_target_qty"] - detail["inventory_qty"])),
        0,
    )
    detail["suggest_outbound_qty"] = np.where(
        (detail["risk_level"] == "高") & (detail["forecast_daily_sales"] > 0),
        np.floor(np.maximum(0, detail["inventory_qty"] - detail["high_keep_qty"])),
        0,
    )
    detail["suggest_replenish_qty"] = np.where(
        (detail["risk_level"] == "低") | (detail["out_of_stock"] == "是"),
        detail["need_qty"],
        0,
    )
    detail["suggest_outbound_qty"] = detail["suggest_outbound_qty"].astype(int)
    detail["suggest_replenish_qty"] = detail["suggest_replenish_qty"].astype(int)

    # Rename columns to Chinese for output
    detail_out = detail.rename(columns={
        "store": "门店名称",
        "brand": "品牌",
        "barcode": "商品条码",
        "product": "商品名称",
        "daily_sales_3m_mtd": "近三月+本月迄今平均日销",
        "daily_sales_30d": "近30天平均日销售",
        "inventory_qty": "库存数量",
        "out_of_stock": "缺货",
        "risk_level": "风险等级",
        "inventory_sales_ratio": "库存/销售比",
        "turnover_rate": "库存周转率",
        "turnover_days": "库存周转天数",
        "suggest_outbound_qty": "建议调出数量",
        "suggest_replenish_qty": "建议补货数量",
    })
    detail_out["商品条码"] = detail_out["商品条码"].astype(str)
    detail_out = detail_out[[
        "门店名称",
        "品牌",
        "商品条码",
        "商品名称",
        "近三月+本月迄今平均日销",
        "近30天平均日销售",
        "库存数量",
        "缺货",
        "风险等级",
        "库存/销售比",
        "库存周转率",
        "库存周转天数",
        "建议调出数量",
        "建议补货数量",
    ]]

    store_summary_out = store_summary.rename(columns={
        "store": "门店名称",
        "daily_sales_3m_mtd": "近三月+本月迄今平均日销",
        "daily_sales_30d": "近30天平均日销售",
        "forecast_daily_sales": "预测平均日销(季节模式后)",
        "inventory_qty": "库存数量",
        "risk_level": "风险等级",
        "inventory_sales_ratio": "库存/销售比",
        "turnover_rate": "库存周转率",
        "turnover_days": "库存周转天数",
    })
    store_summary_out = store_summary_out[[
        "门店名称",
        "近三月+本月迄今平均日销",
        "近30天平均日销售",
        "预测平均日销(季节模式后)",
        "库存数量",
        "库存/销售比",
        "库存周转率",
        "库存周转天数",
        "风险等级",
    ]]

    brand_summary_out = brand_summary.rename(columns={
        "brand": "品牌",
        "daily_sales_3m_mtd": "近三月+本月迄今平均日销",
        "daily_sales_30d": "近30天平均日销售",
        "forecast_daily_sales": "预测平均日销(季节模式后)",
        "inventory_qty": "库存数量",
        "risk_level": "风险等级",
        "inventory_sales_ratio": "库存/销售比",
        "turnover_rate": "库存周转率",
        "turnover_days": "库存周转天数",
    })
    brand_summary_out = brand_summary_out[[
        "品牌",
        "近三月+本月迄今平均日销",
        "近30天平均日销售",
        "预测平均日销(季节模式后)",
        "库存数量",
        "库存/销售比",
        "库存周转率",
        "库存周转天数",
        "风险等级",
    ]]

    missing_out = missing_sales[[
        "store",
        "brand",
        "barcode",
        "product",
        "daily_sales_3m_mtd",
        "daily_sales_30d",
    ]].rename(columns={
        "store": "门店名称",
        "brand": "品牌",
        "barcode": "商品条码",
        "product": "商品名称",
        "daily_sales_3m_mtd": "近三月+本月迄今平均日销",
        "daily_sales_30d": "近30天平均日销售",
    })
    missing_out["商品条码"] = missing_out["商品条码"].astype(str)
    replenish_lookup_exact = (
        detail_out[["门店名称", "品牌", "商品条码", "商品名称", "建议补货数量"]]
        .groupby(["门店名称", "品牌", "商品条码", "商品名称"], as_index=False)["建议补货数量"]
        .max()
    )
    missing_out = missing_out.merge(
        replenish_lookup_exact,
        on=["门店名称", "品牌", "商品条码", "商品名称"],
        how="left",
    )
    replenish_lookup_fallback = (
        detail_out[["门店名称", "品牌", "商品名称", "建议补货数量"]]
        .groupby(["门店名称", "品牌", "商品名称"], as_index=False)["建议补货数量"]
        .max()
        .rename(columns={"建议补货数量": "建议补货数量_fallback"})
    )
    missing_out = missing_out.merge(
        replenish_lookup_fallback,
        on=["门店名称", "品牌", "商品名称"],
        how="left",
    )
    missing_out["建议补货数量"] = (
        missing_out["建议补货数量"].fillna(missing_out["建议补货数量_fallback"]).fillna(0).astype(int)
    )
    missing_out = missing_out.drop(columns=["建议补货数量_fallback"])
    missing_out = missing_out[[
        "门店名称",
        "品牌",
        "商品条码",
        "商品名称",
        "近三月+本月迄今平均日销",
        "近30天平均日销售",
        "建议补货数量",
    ]]

    replenish_out = detail_out[detail_out["建议补货数量"] > 0].copy()
    replenish_out = replenish_out[[
        "门店名称",
        "品牌",
        "商品条码",
        "商品名称",
        "近三月+本月迄今平均日销",
        "近30天平均日销售",
        "库存数量",
        "缺货",
        "风险等级",
        "建议补货数量",
    ]]

    transfer_out = detail_out[detail_out["建议调出数量"] > 0].copy()
    transfer_out = transfer_out[[
        "门店名称",
        "品牌",
        "商品条码",
        "商品名称",
        "近三月+本月迄今平均日销",
        "近30天平均日销售",
        "库存数量",
        "风险等级",
        "建议调出数量",
    ]]

    factor_by_barcode = (
        carton_factor_df[carton_factor_df["商品条码"].notna()][["商品条码", "装箱数（因子）"]]
        .drop_duplicates(subset=["商品条码"], keep="first")
    )
    factor_by_product = (
        carton_factor_df[["商品名称", "装箱数（因子）"]]
        .drop_duplicates(subset=["商品名称"], keep="first")
    )

    def attach_factor_and_case_count(
        df: pd.DataFrame,
        qty_col: str,
        case_col: Optional[str] = None,
        case_with_unit: bool = False,
        include_factor_col: bool = True,
    ) -> pd.DataFrame:
        out = df.copy()
        out["商品条码"] = out["商品条码"].apply(normalize_barcode_value)
        out = out.merge(factor_by_barcode, on="商品条码", how="left")
        fallback = out["装箱数（因子）"].isna()
        if fallback.any():
            merged = out.loc[fallback, ["商品名称"]].merge(
                factor_by_product,
                on="商品名称",
                how="left",
                suffixes=("", "_fallback"),
            )
            out.loc[fallback, "装箱数（因子）"] = merged["装箱数（因子）_fallback"].values

        out["装箱数（因子）"] = pd.to_numeric(out["装箱数（因子）"], errors="coerce")
        valid_factor = out["装箱数（因子）"] > 0
        if case_col:
            out[case_col] = compute_case_counts(out[qty_col], out["装箱数（因子）"], use_peak_mode)
            if case_with_unit:
                out[case_col] = out[case_col].apply(lambda x: f"{int(x)}件" if pd.notna(x) else "")

        out["装箱数（因子）"] = np.where(valid_factor, out["装箱数（因子）"].astype(int), np.nan)

        cols = out.columns.tolist()
        if "装箱数（因子）" in cols:
            cols.remove("装箱数（因子）")
            if include_factor_col:
                product_idx = cols.index("商品名称")
                cols.insert(product_idx + 1, "装箱数（因子）")
        if case_col and case_col in cols:
            cols.remove(case_col)
            cols.append(case_col)
        return out[cols]

    missing_out = attach_factor_and_case_count(
        missing_out,
        "建议补货数量",
        case_col="建议补货箱数",
        case_with_unit=True,
        include_factor_col=False,
    )

    replenish_out = attach_factor_and_case_count(
        replenish_out,
        "建议补货数量",
        case_col="建议补货箱数",
        case_with_unit=True,
        include_factor_col=True,
    )
    transfer_out = attach_factor_and_case_count(
        transfer_out,
        "建议调出数量",
        case_col=None,
        include_factor_col=True,
    )
    if "建议调货箱数" in transfer_out.columns:
        transfer_out = transfer_out.drop(columns=["建议调货箱数"])

    # Summary sheet
    summary_rows = [
        ["风险等级-高", int((detail_out["风险等级"] == "高").sum())],
        ["风险等级-中", int((detail_out["风险等级"] == "中").sum())],
        ["风险等级-低", int((detail_out["风险等级"] == "低").sum())],
        ["缺货/库存缺失SKU数", int(len(missing_out))],
        ["库存总量", float(detail_out["库存数量"].sum())],
        ["近三月+本月迄今平均日销总量", round(float(detail_out["近三月+本月迄今平均日销"].sum()), 1)],
        ["近30天平均日销售总量", round(float(detail_out["近30天平均日销售"].sum()), 1)],
        ["预测平均日销总量(季节模式后)", round(float(detail["forecast_daily_sales"].sum()), 1)],
    ]
    summary_out = pd.DataFrame(summary_rows, columns=["指标", "数值"])
    status_rows = [
        ["季节模式", "旺季(取高值)" if use_peak_mode else "淡季(取低值)"],
        ["3M+MTD窗口有效", "是" if has_mtd_window_data else "否"],
        ["30天窗口有效", "是" if has_recent_window_data else "否"],
        ["销售无效日期行数", invalid_sales_date_rows],
        [
            "建议补货清单缺失装箱因子行数",
            int((replenish_out["装箱数（因子）"].isna()).sum()) if "装箱数（因子）" in replenish_out.columns else 0,
        ],
        [
            "建议调货清单缺失装箱因子行数",
            int((transfer_out["装箱数（因子）"].isna()).sum()) if "装箱数（因子）" in transfer_out.columns else 0,
        ],
        [
            "窗口数据状态",
            "正常" if (has_mtd_window_data and has_recent_window_data)
            else f"警告: 3M+MTD有效={has_mtd_window_data}, 30天有效={has_recent_window_data}",
        ],
    ]
    status_out = pd.DataFrame(status_rows, columns=["状态项", "值"])

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        detail_out.to_excel(writer, sheet_name="明细", index=False)
        store_summary_out.to_excel(writer, sheet_name="门店汇总", index=False)
        brand_summary_out.to_excel(writer, sheet_name="品牌汇总", index=False)
        missing_out.to_excel(writer, sheet_name="缺货清单", index=False)
        replenish_out.to_excel(writer, sheet_name="建议补货清单", index=False)
        transfer_out.to_excel(writer, sheet_name="建议调货清单", index=False)
        summary_out.to_excel(writer, sheet_name="汇总", index=False)
        status_out.to_excel(writer, sheet_name="运行状态", index=False)

    # Apply styling and merging
    wb = load_workbook(output_file)

    header_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    risk_fills = {
        "高": PatternFill(start_color="F44336", end_color="F44336", fill_type="solid"),
        "中": PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid"),
        "低": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
    }
    out_of_stock_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")

    def style_sheet(ws, sheet_name: str):
        ws.insert_rows(1)
        title = f"{display_name} | 库存日期：{inventory_date}"
        ws.cell(row=1, column=1, value=title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=1).alignment = center

        ws.freeze_panes = "E3" if sheet_name == "明细" else "A3"
        last_col = ws.max_column
        last_col_letter = ws.cell(row=2, column=last_col).column_letter
        ws.auto_filter.ref = f"A2:{last_col_letter}{ws.max_row}"
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # Auto width based on cell content (with caps for readability)
        headers = [c.value for c in ws[2]]
        for col in ws.columns:
            max_len = 0
            col_letter = ws.cell(row=2, column=col[0].column).column_letter
            header = ws.cell(row=2, column=col[0].column).value
            for cell in col:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            auto_width = max_len + 2
            if header in PREFERRED_WIDTHS:
                ws.column_dimensions[col_letter].width = max(PREFERRED_WIDTHS[header], min(auto_width, 40))
            else:
                ws.column_dimensions[col_letter].width = min(auto_width, 40)

        header_to_idx = {name: idx + 1 for idx, name in enumerate(headers)}
        wrap_columns = {"门店名称", "商品名称"}
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            for col_idx, cell in enumerate(row, start=1):
                header = headers[col_idx - 1]
                if header in wrap_columns:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif isinstance(cell.value, (int, float)):
                    cell.alignment = center
                else:
                    cell.alignment = left

                if header in NUMBER_FORMATS:
                    cell.number_format = NUMBER_FORMATS[header]

            risk_idx = header_to_idx.get("风险等级")
            if risk_idx is not None:
                risk_cell = row[risk_idx - 1]
                risk_fill = risk_fills.get(risk_cell.value)
                if risk_fill:
                    risk_cell.fill = risk_fill
                    risk_cell.font = Font(color="FFFFFF", bold=True)
                    risk_cell.alignment = center

            barcode_idx = header_to_idx.get("商品条码")
            if barcode_idx is not None:
                row[barcode_idx - 1].number_format = "@"

            avg_idx = header_to_idx.get("近三月+本月迄今平均日销")
            inv_idx = header_to_idx.get("库存数量")
            if avg_idx is not None and inv_idx is not None:
                avg_val = row[avg_idx - 1].value or 0
                inv_val = row[inv_idx - 1].value or 0
                if avg_val > 0 and inv_val == 0:
                    for cell in row:
                        cell.fill = out_of_stock_fill
                        cell.font = Font(color="FFFFFF", bold=True)

        if sheet_name == "汇总" and "指标" in header_to_idx and "数值" in header_to_idx:
            metric_idx = header_to_idx["指标"] - 1
            value_idx = header_to_idx["数值"] - 1
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                metric_name = row[metric_idx].value
                if metric_name in SUMMARY_ONE_DECIMAL_METRICS and isinstance(row[value_idx].value, (int, float)):
                    row[value_idx].number_format = "0.0"

    for sheet in ["明细", "门店汇总", "品牌汇总", "缺货清单", "建议补货清单", "建议调货清单", "汇总", "运行状态"]:
        style_sheet(wb[sheet], sheet)

    # Merge same store in Detail sheet
    ws_detail = wb["明细"]
    headers = [c.value for c in ws_detail[2]]
    if "门店名称" in headers:
        store_idx = headers.index("门店名称") + 1
        start_row = 3
        current = ws_detail.cell(row=3, column=store_idx).value
        for row in range(4, ws_detail.max_row + 2):
            value = ws_detail.cell(row=row, column=store_idx).value if row <= ws_detail.max_row else None
            if value != current:
                end_row = row - 1
                if end_row > start_row:
                    ws_detail.merge_cells(start_row=start_row, start_column=store_idx, end_row=end_row, end_column=store_idx)
                    ws_detail.cell(row=start_row, column=store_idx).alignment = center
                start_row = row
                current = value

    # Apply borders after merges for all active cells
    from openpyxl.styles import Border, Side
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border

    wb.save(output_file)

    print(f"[{display_name}] Report saved: {output_file}")
    return {
        "system_id": system_id,
        "display_name": display_name,
        "status": "SUCCESS",
        "message": "",
        "output_file": str(output_file),
        "detail_rows": int(len(detail_out)),
        "missing_rows": int(len(missing_out)),
    }

def run_batch(global_config: Dict[str, Any]) -> int:
    validate_batch_config(global_config)
    batch_cfg = global_config["batch"]
    continue_on_error = bool(batch_cfg.get("continue_on_error", True))
    summary_output_file = Path(batch_cfg.get("summary_output_file", "./reports/batch_run_summary.xlsx"))
    if not summary_output_file.is_absolute():
        summary_output_file = (BASE_DIR / summary_output_file).resolve()
    summary_output_file.parent.mkdir(parents=True, exist_ok=True)

    records: List[Dict[str, Any]] = []
    failure_count = 0
    for system in batch_cfg.get("systems", []):
        start_time = datetime.now()
        enabled = bool(system.get("enabled", True))
        display_name = str(system.get("display_name", system.get("system_id", "unknown")))
        system_id = str(system.get("system_id", "")).strip() or display_name
        data_subdir = str(system.get("data_subdir", "")).strip()
        expected_output_file = resolve_expected_output_for_status(system, display_name)
        merged_config: Optional[Dict[str, Any]] = None
        try:
            merged_config = build_system_config(system, global_config)
            display_name = merged_config["display_name"]
            system_id = merged_config["system_id"]
            expected_output_file = resolve_expected_output_for_status(merged_config, display_name)
        except Exception:
            # Keep original fallback values for summary when config merge itself fails.
            merged_config = None

        if not enabled:
            records.append(
                {
                    "system_id": system_id,
                    "display_name": display_name,
                    "enabled": False,
                    "data_subdir": data_subdir,
                    "status": "SKIPPED",
                    "message": "disabled",
                    "output_file": expected_output_file,
                    "duration_sec": 0.0,
                    "generated_at": datetime.now().isoformat(timespec="seconds"),
                    "detail_rows": 0,
                    "missing_rows": 0,
                }
            )
            print(f"[{display_name}] Skipped: disabled")
            continue
        try:
            if merged_config is None:
                merged_config = build_system_config(system, global_config)
                display_name = merged_config["display_name"]
                system_id = merged_config["system_id"]
                expected_output_file = resolve_expected_output_for_status(merged_config, display_name)
            print(f"[{display_name}] Start generating report...")
            record = generate_report_for_system(merged_config, global_config)
        except Exception as exc:  # noqa: BLE001
            failure_count += 1
            record = {
                "system_id": system_id,
                "display_name": display_name,
                "enabled": True,
                "data_subdir": data_subdir,
                "status": "FAILED",
                "message": str(exc),
                "output_file": expected_output_file,
                "detail_rows": 0,
                "missing_rows": 0,
            }
            print(f"[{display_name}] Failed: {exc}")
            if not continue_on_error:
                record["duration_sec"] = round((datetime.now() - start_time).total_seconds(), 3)
                record["generated_at"] = datetime.now().isoformat(timespec="seconds")
                records.append(record)
                break

        record["enabled"] = True
        record["data_subdir"] = data_subdir
        record["duration_sec"] = round((datetime.now() - start_time).total_seconds(), 3)
        record["generated_at"] = datetime.now().isoformat(timespec="seconds")
        records.append(record)

    summary_df = pd.DataFrame(
        records,
        columns=[
            "system_id",
            "display_name",
            "enabled",
            "data_subdir",
            "status",
            "message",
            "output_file",
            "duration_sec",
            "generated_at",
            "detail_rows",
            "missing_rows",
        ],
    )
    summary_df.to_excel(summary_output_file, index=False)
    print(f"Batch summary saved: {summary_output_file}")
    return failure_count


def main() -> int:
    config = load_config()
    run_mode = str(config.get("run_mode", "single")).lower()
    if run_mode == "batch":
        failures = run_batch(config)
        return 1 if failures > 0 else 0

    generate_report_for_system(config, config)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
